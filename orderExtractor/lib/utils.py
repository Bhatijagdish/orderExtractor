from typing import List

import openpyxl.utils.cell
from openpyxl.styles import Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo

CURRENCY_FORMAT = "₹ #,##0.00"


def create_sales_snapshot_dict(channelNames, totalSales, orderStatuses,
                               shippingFees, serviceTaxes,
                               settlementAmounts, channelFees) -> dict:
    data = {}
    for i in range(len(channelNames)):
        channel_name = trim_channel_name(channelNames[i])
        if channel_name not in data.keys():
            data[channel_name] = [0] * 6

        currentValues = data[channel_name]
        currentValues[0] = currentValues[0] + float(totalSales[i])

        if "cancel" in orderStatuses[i].lower() and "return" not in orderStatuses[i].lower():
            currentValues[1] = currentValues[1] + float(totalSales[i])

        if "return" in orderStatuses[i].lower() and float(shippingFees[i]) != 0:
            currentValues[3] = currentValues[3] + float(totalSales[i])

        if "return" in orderStatuses[i].lower() and float(shippingFees[i]) == 0:
            currentValues[2] = currentValues[2] + float(totalSales[i])

        currentValues[1] = currentValues[1]
        currentValues[2] = currentValues[2]
        currentValues[3] = currentValues[3]
        currentValues[4] = currentValues[4] + float(serviceTaxes[i]) + float(channelFees[i])
        currentValues[5] = currentValues[5] + float(settlementAmounts[i])
        data[channel_name] = currentValues

    for key, values in data.items():
        data[key] = [values[0], values[1], values[2], values[3],
                     (values[0] - values[1] - values[2] - values[3]),
                     values[4], values[5]]
    data["Total"] = [sum(values) for values in zip(*data.values())]

    return data


def trim_channel_name(channel_name: str):
    if "-" in channel_name:
        return channel_name.split("-")[1].strip()
    else:
        return channel_name.strip()


def getTableCoordinates(table_name: str, sheet) -> str:
    if table_name in sheet.tables:
        table_range = sheet.tables[table_name].ref
        return table_range


def get_table_last_row(table_range: str):
    _, end = table_range.split(":")
    last_row = float(''.join(filter(str.isdigit, end)))
    return last_row


def paste_data(dict_data: dict, sheet, formatRequired: bool = True) -> None:
    # Loop to update the information to the data sheet
    for row_index, (key, values) in enumerate(dict_data.items(), start=2):
        sheet.cell(row=row_index, column=1, value=key)
        for col_idx, value in enumerate(values, start=2):
            cell = sheet.cell(row=row_index, column=col_idx, value=value)
            if formatRequired:
                cell.number_format = CURRENCY_FORMAT


def apply_currency_format(sheet, col_idx: int) -> None:
    last_row = sheet.max_row
    last_col = sheet.max_column
    for col in range(col_idx, last_col + 1):
        for row in range(2, last_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = CURRENCY_FORMAT


def get_thick_border():
    thick_border = Side(border_style="thick", color="000000")
    return Border(top=thick_border, left=thick_border, bottom=thick_border, right=thick_border)


def create_table(table_name: str, sheet, table_range: str = None) -> None:
    if table_range is not None:
        table_ref = table_range
    else:
        table_ref = f"A1:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}"
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)


def read_table_using_table_name(table_name: str, sheet):
    table = sheet.tables[table_name]

    start, end = table.ref.split(":")
    start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start)
    end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end)

    table_values = []
    for row in range(start_row + 1, end_row + 1):
        row_values = []
        for col in openpyxl.utils.cols_from_range(f"{start_col}{row}:{end_col}{row}")[0]:
            cell = sheet[col]
            row_values.append(cell.value)
        table_values.append(row_values)

    return table_values


def create_channel_dict(productSkuLists: List,
                        trimChannelName: str,
                        channelNames: List,
                        taxRates: List,
                        totalSales: List,
                        orderStatuses: List,
                        shippingFees: List,
                        serviceTaxes: List,
                        quantities: List,
                        channelFees: List,
                        product_cost_dict: dict,
                        ad_expenses: float = 0,
                        labour_charges: float = 0) -> dict:
    """
    trimChannelName is the string of selected channel "Amazon, Flipkart etc"
    labour charges can be an amount that is Optional
    product_cost_dict will be static or dynamic but comes from first execution
    ad_expenses comes from a dictionary created after first execution but can be an optional amount
    """
    data = {}
    for i in range(len(productSkuLists)):
        sku_name = productSkuLists[i]
        channel_name = trim_channel_name(channelNames[i])
        if channel_name == trimChannelName:

            if sku_name not in data.keys():
                data[sku_name] = [0] * 27

            currentValues = data[sku_name]

            if "delivered" in orderStatuses[i].lower():
                currentValues[10] = currentValues[10] + float(totalSales[i])

            if "cancel" in orderStatuses[i].lower() and "return" not in orderStatuses[i].lower():
                currentValues[3] = currentValues[3] + float(quantities[i])

            if "return" in orderStatuses[i].lower() and float(shippingFees[i]) == 0:
                currentValues[5] = currentValues[5] + float(quantities[i])

            if "return" in orderStatuses[i].lower() and float(shippingFees[i]) != 0:
                currentValues[6] = currentValues[6] + float(quantities[i])

            currentValues[0] = taxRates[i]
            currentValues[1] = currentValues[1] + 1
            currentValues[2] = currentValues[2] + float(quantities[i])
            currentValues[3] = currentValues[3]
            # currentValues[4] = round(currentValues[3] / currentValues[2] * 100, 2)
            currentValues[5] = currentValues[5]
            currentValues[6] = currentValues[6]
            # currentValues[7] = round((currentValues[5] + currentValues[6]) / currentValues[2] * 100, 2)
            # currentValues[8] = currentValues[2] - currentValues[3] - currentValues[5] - currentValues[6]
            currentValues[10] = currentValues[10]
            currentValues[11] = currentValues[11] + float(serviceTaxes[i]) + float(channelFees[i])

            data[sku_name] = currentValues

    # sum of second index of each sku data
    total_orders = sum(value[1] for value in data.values())

    for key, values in data.items():
        product_cost = product_cost_dict[key]

        cancel_ratio = round(values[3] / values[2] * 100, 2)
        return_ratio = round((values[5] + values[6]) / values[2] * 100, 2)
        total_return_ratio = round((values[3] + values[5] + values[6]) / values[2] * 100, 2)
        fulfilled_pcs = values[2] - values[3] - values[5] - values[6]
        gross_amount = values[10] - values[11]
        tcs_tds = round((values[10] - values[11]) * 2 / 100, 2)
        gst_amt = round(values[10] - (values[10] / (100 + values[0]) * 100), 2)

        # input for labour_charges
        labour_charge = values[1] * labour_charges
        gross_amount_nxt = round(gross_amount - tcs_tds - gst_amt - labour_charge, 2)
        gst_tcs_tds = round(values[11] - values[11] / (100 + values[0]) * 100 + tcs_tds, 2)
        product_gst_input = round(product_cost - product_cost / (100 + values[0]) * 100, 2) * fulfilled_pcs

        amt_realised = gross_amount_nxt + gst_tcs_tds + product_gst_input
        ad_expense = round(values[1] / total_orders * ad_expenses, 2)
        net_realised_amount = amt_realised - ad_expense
        per_piece = round(net_realised_amount / fulfilled_pcs, 2) if total_return_ratio != 100 else 0
        net_profit_per_piece = round(per_piece - product_cost, 2)
        total_net_profit = round((- values[11] - labour_charge - ad_expense if total_return_ratio == 100
                                  else net_profit_per_piece * fulfilled_pcs), 2)
        average_selling_price = round(values[10] / fulfilled_pcs, 2) if total_return_ratio != 100 else 0

        new_value = values

        new_value[4] = cancel_ratio
        new_value[7] = return_ratio
        new_value[8] = total_return_ratio
        new_value[9] = fulfilled_pcs
        new_value[11] = new_value[11]
        new_value[12] = gross_amount
        new_value[13] = tcs_tds
        new_value[14] = gst_amt
        new_value[15] = labour_charge
        new_value[16] = gross_amount_nxt
        new_value[17] = gst_tcs_tds
        new_value[18] = product_gst_input
        new_value[19] = amt_realised
        new_value[20] = ad_expense
        new_value[21] = net_realised_amount
        new_value[22] = per_piece
        new_value[23] = product_cost
        new_value[24] = net_profit_per_piece
        new_value[25] = total_net_profit
        new_value[26] = average_selling_price

        data[key] = new_value

    last_row = [None] * 27
    last_row[1] = total_orders
    last_row[2] = sum(value[2] for value in data.values())
    last_row[3] = sum(value[3] for value in data.values())
    last_row[4] = round(last_row[3] / last_row[2] * 100, 2)
    last_row[5] = sum(value[5] for value in data.values())
    last_row[6] = sum(value[6] for value in data.values())
    last_row[7] = round((last_row[5] + last_row[6]) / last_row[2] * 100, 2)
    last_row[8] = round((last_row[3] + last_row[5] + last_row[6]) / last_row[2] * 100, 2)
    last_row[9] = sum(value[9] for value in data.values())
    last_row[10] = sum(value[10] for value in data.values())
    last_row[11] = sum(value[11] for value in data.values())
    last_row[15] = sum(value[15] for value in data.values())
    last_row[18] = sum(value[18] for value in data.values())
    last_row[20] = sum(value[20] for value in data.values())
    last_row[21] = sum(value[21] for value in data.values())
    last_row[23] = "Net Profit"
    last_row[25] = sum(value[25] for value in data.values())

    data["Total"] = last_row

    return data


def get_ad_expenses_value(workbook, channel_names: List) -> dict:
    # last row that contains channel data considering starting from the row 1 as headers
    last_row = len(channel_names) + 1

    ad_expenses = {}
    sheet = workbook.worksheets[1]
    # iterate up to the expected last row only
    for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=10):
        ad_expenses[row[0]] = row[9]

    return ad_expenses



