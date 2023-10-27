import os
import sys
from typing import List
import openpyxl
from tkinter import messagebox
from openpyxl.styles.table import TableStyle
from openpyxl.worksheet.table import Table

from lib.utils import read_table_using_table_name


class ReadFile:

    def __init__(self, filepath: str):
        self.file_path = filepath
        self.orders_data = None
        self.workbook = None
        self.input_sheet = None
        self.output_sheet = None

    # def read_orders_data(self, sheetname: Any = None):
    #     try:
    #         if self.file_path.endswith(".csv"):
    #             self.orders_data = pd.read_csv(self.file_path)
    #         elif self.file_path.endswith(".xlsx") or self.file_path.endswith(".xls")
    #         or self.file_path.endswith(".xl"):
    #             self.orders_data = pd.read_excel(self.file_path,
    #                                              index_col=0,
    #                                              sheet_name=sheetname)
    #         else:
    #             messagebox.showerror("Unsupported File Type", "File Type not supported, please select "
    #                                                           "the valid csv or excel file")
    #     except:
    #         messagebox.showerror("Unsupported File Type", "File Type not supported, please select "
    #                                                       "the valid csv or excel file")

    def get_sheet_using_sheet_name(self, sheetname: int | str):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
        except:
            messagebox.showerror("Unsupported File Type", "File Type not supported, please select "
                                                          "the valid csv or excel file")
            sys.exit(0)
        try:
            if type(sheetname) == int:
                self.input_sheet = self.workbook.worksheets[sheetname]
            else:
                self.input_sheet = self.workbook[sheetname]
        except:
            messagebox.showerror("Not Found", f"Unable to find a sheet with the given name: {sheetname}")
            sys.exit(0)

    def get_rows_data(self, sheet):
        sheet.iter_rows(values_only=True)

    # def add_cell_value(self, sheet: str, cell: str, value: str):
    #     try:
    #         self.orders_data.at[cell] = value
    #         self.orders_data.to_excel(self.file_path, sheetname, index=False)
    #     except:
    #         messagebox.showerror("Error", f"Unable to update the value: {value} "
    #                                       f"at position {cell} in sheet {sheetname}")

    def add_cell_value_pyxl(self, sheetname: str, cell: str, value: str):
        try:
            sheet = self.workbook[sheetname]
            sheet[cell] = value
            self.workbook.save(self.file_path)
        except:
            messagebox.showerror("Error", f"Unable to update the value: {value} "
                                          f"at position {cell} in sheet {sheetname}")
            sys.exit(0)

    def update_net_realisation_value_by_channel_name(self, sheet_name: str, channel_name: str,
                                                     net_realisation_amount: int) -> None:
        try:
            sheet = self.workbook[sheet_name]
            for row in sheet.iterrows(min_col=1, max_col=9, values_only=False):
                if row[0].value == channel_name:
                    row[8].value = net_realisation_amount
                    break
        except:
            pass

    def create_new_sheet(self, sheet_name: str) -> None:
        self.output_sheet = self.workbook.create_sheet(sheet_name)

    def get_column_data_by_header(self, header_name: str):
        for col in self.input_sheet.iter_cols(1, self.input_sheet.max_column):
            if col[0].value == header_name:
                return [cell.value for cell in col[1:]]
        return []

    def create_table(self, sheet_name: str, table_data: List[List[str]], table_name: str):

        sheet = self.workbook[sheet_name]
        for row in table_data:
            sheet.append(row)

        table = Table(displayName=table_name, ref=sheet.dimensions)

        style = TableStyle(name="TableStyleMedium9"
                           # showFirstColumn=False,
                           # showLastColumn=False,
                           # showRowStripes=True,
                           # showColumnStripes=True
                           )

        table.tableStyleInfo = style
        sheet.add_table(table)

    def save_workbook(self):
        self.workbook.save(self.file_path)

    def adjust_width(self) -> None:
        # Iterate over all columns and adjust their widths
        for column in self.output_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width < 10:
                adjusted_width = 10
            self.output_sheet.column_dimensions[column_letter].width = adjusted_width

    def get_static_sku_data(self):
        data = {}
        for row in self.input_sheet.iter_rows(min_col=1, max_col=2, values_only=True):
            sku, cost = row
            if sku in data.keys():
                file_name = os.path.basename(self.file_path)
                messagebox.showerror("Duplicate SKU", f"SKU value in {file_name} "
                                                      f"is repeating which may create conflicts in the report")
                sys.exit(0)
            if sku not in data.keys():
                data[sku] = cost

        return data

    def get_dynamic_sku_data(self, table_name):
        sku_code_listing = cost_list = []
        for sheetname in self.workbook.sheetnames:
            if "sales snapshot" in sheetname.lower():
                sheet = self.workbook[sheetname]
                row = read_table_using_table_name(table_name, sheet)
                for sku, cost in row:
                    sku_code_listing.append(sku)
                    cost_list.append(cost)
        return sku_code_listing, cost_list
