from typing import List

from openpyxl.styles import Font
from lib.file_selector import select_file
from lib.utils import create_sales_snapshot_dict, paste_data, create_table, get_thick_border, create_channel_dict, \
    apply_currency_format, get_ad_expenses_value
from lib.xl_csv_reader import ReadFile
from orders_data import Orders


def snapshot_creation():
    file_path: str = select_file()
    data = Orders(file_path)
    try:
        month = data.get_month()
        data.create_new_sheet(f"Sales Snapshot for {month} Month")

        first_table_headings = ['List of MarketPlaces', 'GMV', 'Cancelled', 'RTO', 'RTV',
                                'Net Sales', 'MarketPlace Charges', 'Settlement Amount',
                                'Net Realisation Amount', 'Ad Expenses']

        data.output_sheet.append(first_table_headings)

        new_channel_data = create_sales_snapshot_dict(data.get_channel_names(), data.total_sales(),
                                                      data.order_statuses(),
                                                      data.shipping_fee(), data.service_tax(),
                                                      data.settlement_amounts(),
                                                      data.channel_fees())

        paste_data(new_channel_data, data.output_sheet)
        create_table("channelDataTable", data.output_sheet)

        # applying styles to the table

        for cell in data.output_sheet[data.output_sheet.max_row]:
            cell.font = Font(bold=True)
            cell.border = get_thick_border()

        # Creating Listing SKu Table in Sales Snapshot table

        last_row = data.output_sheet.max_row

        data.output_sheet.cell(row=last_row + 4, column=1, value="Listing Sku Code")
        data.output_sheet.cell(row=last_row + 4, column=2, value="Total Cost")

        for index, value in enumerate(set(data.list_sku_code()), start=last_row + 5):
            data.output_sheet.cell(row=index, column=1, value=value)

        create_table("skuListingTable", data.output_sheet, f"A{last_row + 4}:B{data.output_sheet.max_row}")

        data.adjust_width()

        data.save_workbook()
    finally:
        data.workbook.close()


def static_product_cost():
    file_path: str = select_file()
    rd = ReadFile(file_path)
    rd.get_sheet_using_sheet_name(0)
    try:
        data = rd.get_static_sku_data()
    finally:
        rd.workbook.close()
    return data


def channel_sheet_creation(product_cost_dict: dict, ad_expenses_checked: bool, labour_charges: int = 0):
    file_path: str = select_file()
    data = Orders(file_path)
    try:
        channels: List = data.trim_channel_names()
        ad_expenses_dict: dict = {}
        if ad_expenses_checked:
            ad_expenses_dict = get_ad_expenses_value(data.workbook, channels)
        for channel in channels:
            if ad_expenses_checked:
                ad_expenses = ad_expenses_dict[channel]
            else:
                ad_expenses = 0
            data.create_new_sheet(channel)

            headings = ['Items', 'GST %', 'No Of Orders', 'Sold Pieces', 'Cancelled Pieces', 'Cancelled %', 'RTO',
                        'RTV', 'Returns %', 'Total Return %', 'Fulfilled Pieces', 'Total Invest Amount',
                        'Channel Charges', 'Gross Amount Realised', '(-) TCS & TDS', '(-) GST',
                        '(-) Pack & Labour', 'Gross Amt Realised2', '(+) GST Input & TCS & TDS',
                        '(+) Product GST Input', 'Amt Realised', 'Ads Exp', 'Net Amt Realised',
                        'Per Piece', 'Product Cost', 'Net Profit Per Piece', 'Total Net Profit',
                        'Average Selling Price']

            data.output_sheet.append(headings)

            channel_data = create_channel_dict(data.list_sku_code(),
                                               channel,
                                               data.get_channel_names(),
                                               data.tax_rates(),
                                               data.total_sales(),
                                               data.order_statuses(),
                                               data.shipping_fee(),
                                               data.service_tax(),
                                               data.quantities(),
                                               data.channel_fees(),
                                               product_cost_dict,
                                               labour_charges,
                                               ad_expenses)

            total_net_realisation_amount = channel_data["Total"][25]

            # print(channel_data)
            paste_data(channel_data, data.output_sheet, False)
            displayName = channel.replace(" ", "")
            create_table(f"{displayName}_Table", data.output_sheet)

            # Applying styling to the last row of the table
            for cell in data.output_sheet[data.output_sheet.max_row]:
                cell.font = Font(bold=True)
                cell.border = get_thick_border()

            # Adjusting the width of column A
            data.output_sheet.column_dimensions['A'].width = 50
            apply_currency_format(data.output_sheet, 12)
            data.update_net_realisation_value_by_channel_name(channel, channel, total_net_realisation_amount)

        data.save_workbook()
    finally:
        data.workbook.close()

